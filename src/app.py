import os
import streamlit as st
import pandas as pd
import uuid
import time
import cv2
import tempfile
import base64
from typing import List, Any
import torch
torch.classes.__path__ = [os.path.join(torch.__path__[0], torch.classes.__file__)] 
# or simply:
torch.classes.__path__ = []
# Import modules from current codebase
from du_prompts import DETECT_SENSITIVE_PROMPT1, CHECK_SENSITIVE_PROMPT
from markdown import ocr2markdown
from search import find_sensitive_data_bboxes, mask_sensitive_data
from process_ocr import convert_pdf_to_images, load_ocr_model, perform_ocr, save_ocr_results_as_json
from process_llm import call_model_litellm
from process_presidio import load_analyzer, detect, process_excel_results
from utils import merge_pdf, process_llm_results, process_llm_to_list
import openpyxl
from openpyxl.styles import PatternFill

def save_to_excel_with_highlights(df, sensitive_text_list, output_name):
    """
    Saves dataframe to Excel with highlighted sensitive content
    
    :param df: DataFrame to save
    :param sensitive_text_list: List of sensitive text to highlight
    :param output_name: Name for the output file
    :return: Path to the saved Excel file
    """
    # Create output directory if it doesn't exist
    os.makedirs("output", exist_ok=True)
    
    # Define output path
    output_path = f"output/highlighted_sensitive_data_{output_name}.xlsx"
    
    # Save DataFrame to Excel
    df.to_excel(output_path, index=False)
    
    # Open the workbook and get active sheet
    workbook = openpyxl.load_workbook(output_path)
    worksheet = workbook.active
    
    # Define highlight fill
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Check each cell for sensitive text and apply highlighting
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        for col_idx, cell in enumerate(row, start=1):
            if cell.value and isinstance(cell.value, str):
                # Check if any sensitive text is in the cell
                for text in sensitive_text_list:
                    if text.lower() in cell.value.lower():
                        cell.fill = highlight_fill
                        break
    
    # Save the modified workbook
    workbook.save(output_path)
    return output_path

def process_excel_csv(file_path, analyzer):
    """
    Process Excel/CSV file to detect and highlight sensitive information
    
    :param file_path: Path to the Excel/CSV file
    :param analyzer: Presidio analyzer for detecting sensitive information
    :return: Dictionary containing paths to output files
    """
    file_extension = os.path.splitext(file_path)[1].lower()
    file_name = os.path.basename(file_path).split('.')[0]
    
    # Read the file based on extension
    if file_extension == '.csv':
        df = pd.read_csv(file_path)
    else:  # Excel file
        df = pd.read_excel(file_path)
    
    # Convert to markdown for text processing
    markdown = df.to_markdown(index=False)
    
    # Detect sensitive information using Presidio
    results = detect(text=markdown, analyzer=analyzer)
    sensitive_text_list = process_excel_results(results)
    
    # Sort sensitive text by length (descending) to ensure longer matches are processed first
    sorted_sensitive_text = sorted(sensitive_text_list, key=len, reverse=True)
    
    # Save to Excel with highlighting
    excel_path = save_to_excel_with_highlights(df, sorted_sensitive_text, file_name)
    
    return {
        'excel_path': excel_path,
        'sensitive_count': len(sensitive_text_list),
        'sensitive_items': sensitive_text_list
    }

# Initialize necessary directories
os.makedirs("temp", exist_ok=True)
os.makedirs("output", exist_ok=True)
os.makedirs("markdowns", exist_ok=True)

# Global variables to store loaded models
model_storage = {
    'rec_model': None,
    'det_model': None,
    'analyzer': None
}

# Function to load models
@st.cache_resource
def load_models():
    if not model_storage['rec_model'] or not model_storage['det_model']:
        with st.spinner('Loading OCR model...'):
            model_storage['rec_model'], model_storage['det_model'] = load_ocr_model()
    
    if not model_storage['analyzer']:
        with st.spinner('Loading Presidio analyzer...'):
            model_storage['analyzer'] = load_analyzer()
    
    return model_storage['rec_model'], model_storage['det_model'], model_storage['analyzer']

# Function to detect sensitive information using LLM
def llm_detection(text: str, llm_model: str = 'openai', 
                 temperature: float = 0, key_information: str = None) -> List[str]:
    detect_sensitive_prompt = DETECT_SENSITIVE_PROMPT1.format(
        document=text, 
        key_infomation=key_information if key_information else ""
    )
    
    llm_results = call_model_litellm(
        detect_sensitive_prompt,
        model_name=llm_model,
        temparature=temperature
    )
    print("LLM results:", llm_results)
    return process_llm_results(llm_results)

# Function to detect sensitive information using Presidio
def presidio_detection(text: str, analyzer) -> List[str]:
    results = detect(text, analyzer)
    print("Presidio results:", results)
    return results

# Function for combined sensitive information detection
def combined_detection(text: str, analyzer, llm_model: str = 'openai', 
                      temperature: float = 0, key_information: str = None, 
                      detection_method: str = "combined") -> List[str]:
    if detection_method == "llm":
        return llm_detection(text, llm_model, temperature, key_information)
    
    elif detection_method == "presidio":
        return presidio_detection(text, analyzer)
    
    else:  # combined
        presidio_results = presidio_detection(text, analyzer)
        llm_results = llm_detection(text, llm_model, temperature)
        # Combine and remove duplicates
        return list(set(presidio_results + llm_results))

# Function to process document
def process_document(file_path: str, detection_method: str = "combined", 
                    llm_model: str = 'openai', temp_dir: str = None, 
                    progress_callback=None):
    # Load models
    rec_model, det_model, analyzer = load_models()
    
    # Create temp directory if not provided
    if not temp_dir:
        temp_dir = os.path.join("temp", str(uuid.uuid4()))
        os.makedirs(temp_dir, exist_ok=True)
    
    # Determine file type (PDF or image)
    file_extension = os.path.splitext(file_path)[1].lower()
    is_image = file_extension in ['.jpg', '.jpeg', '.png', '.bmp', '.tiff']
    
    if is_image:
        # If image, process directly
        image_paths = [file_path]
        name = os.path.basename(file_path).split('.')[0]
        image_extension = file_extension  # Use original image extension
    else:
        # If PDF, convert to images
        image_paths, name = convert_pdf_to_images(file_path)
        # Use .png for images extracted from PDF
        image_extension = '.png'  # or '.jpg' depending on format from convert_pdf_to_images
    
    # Setup output directory
    masked_dir = os.path.join(temp_dir, "masked_images")
    os.makedirs(masked_dir, exist_ok=True)
    
    masked_output_path = os.path.join(masked_dir, name)
    os.makedirs(masked_output_path, exist_ok=True)
    
    # Process each page/image
    list_masked_images_path = []
    total_pages = len(image_paths)
    
    for index, image_path in enumerate(image_paths):
        # Update progress
        if progress_callback:
            progress_callback((index + 1) / total_pages)
        
        # Perform OCR
        ocr_results = perform_ocr(image_path, rec_model=rec_model, det_model=det_model)
        
        # Save OCR results
        output_file = os.path.join(temp_dir, f"page_{index+1}_ocr.json")
        result_list = save_ocr_results_as_json(ocr_results, output_file)
        
        # Convert to markdown for text processing
        markdown = ocr2markdown(result_list)
        
        # Detect sensitive information
        results = combined_detection(
            text=markdown,
            analyzer=analyzer,
            llm_model=llm_model,
            temperature=0,
            detection_method=detection_method
        )
        sensitive_data = " ".join(results)
        check_sensitive_prompt = CHECK_SENSITIVE_PROMPT.format(sensitive_data=sensitive_data)
        results_checked = call_model_litellm(check_sensitive_prompt,model_name=llm_model,temparature=0)
        results_checked = process_llm_to_list(results_checked)

        # Save sensitive information to file
        result_path = os.path.join(temp_dir, f"page_{index+1}_result.txt")
        with open(result_path, 'w') as f:
            for item in results_checked:
                f.write("%s\n" % item)
        
        # Find bounding boxes and mask sensitive information
        sensitive_bboxes = find_sensitive_data_bboxes(result_list, result_path)
        masked_img = mask_sensitive_data(image_path, sensitive_bboxes)
        
        # Use image_extension instead of file_extension
        masked_img_to_save_path = os.path.join(masked_output_path, f"masked_page_{index+1}{image_extension}")
        try:
            print(f"Saving masked image to: {masked_img_to_save_path}")
            cv2.imwrite(masked_img_to_save_path, masked_img)
        except Exception as e:
            print(f"Error saving masked image: {str(e)}")

        list_masked_images_path.append(masked_img_to_save_path)
    
    # Determine output type based on input
    if is_image:
        # If input is image, return processed image path
        final_output_path = list_masked_images_path[0]
        # Create a copy in the output directory
        output_image_path = os.path.join("output", f"{os.path.basename(temp_dir)}_masked{file_extension}")
        import shutil
        shutil.copy2(final_output_path, output_image_path)
        return output_image_path
    else:
        # If input is PDF, create final PDF
        pdf_final_path = os.path.join("output", f"{os.path.basename(temp_dir)}_masked.pdf")
        merge_pdf(list_masked_images_path, pdf_final_path)
        return pdf_final_path

# Function to display PDF in Streamlit
def display_pdf(file_path):
    with open(file_path, "rb") as f:
        base64_pdf = base64.b64encode(f.read()).decode('utf-8')
    
    pdf_display = f"""
    <iframe src="data:application/pdf;base64,{base64_pdf}" width="100%" height="600" type="application/pdf"></iframe>
    """
    st.markdown(pdf_display, unsafe_allow_html=True)

# Main page of the Streamlit application
def main():
    st.set_page_config(page_title="Automatic Sensitive Information Redaction App", 
                      page_icon="🔒", layout="wide")
    
    st.title("🔒 Automatic Sensitive Information Redaction App")
    st.markdown("""
    Upload a PDF document or image to automatically detect and redact sensitive information such as names, addresses, phone numbers, emails, and other personal data.
    """)
    
    # Sidebar for configuration
    st.sidebar.title("Configuration")
    detection_method = st.sidebar.radio(
        "Select detection method:",
        ["Combined (LLM + Presidio)", "LLM Only", "Presidio Only"],
        index=0
    )
    
    method_map = {
        "Combined (LLM + Presidio)": "combined",
        "LLM Only": "llm",
        "Presidio Only": "presidio"
    }
    
    selected_method = method_map[detection_method]
    
    llm_model = st.sidebar.selectbox(
        "Select LLM model:",
        ["openai", "gemini-2.0-flash","local_llm"],
        index=0,
        disabled=(selected_method == "presidio")
    )
    
    # File upload area
    uploaded_file = st.file_uploader("Upload a PDF document, image, or Excel/CSV file", 
                                    type=["pdf", "jpg", "jpeg", "png", "csv", "xlsx", "xls"])
    
    
    if uploaded_file:
        col1, col2 = st.columns(2)
        
        with col1:
            st.subheader("Original Document")
            
            # Display uploaded file
            file_extension = os.path.splitext(uploaded_file.name)[1].lower()
            
            with tempfile.NamedTemporaryFile(delete=False, suffix=file_extension) as tmp:
                tmp.write(uploaded_file.getvalue())
                temp_path = tmp.name
            
            if file_extension in ['.jpg', '.jpeg', '.png']:
                st.image(temp_path, caption="Uploaded Image")
            elif file_extension in ['.csv', '.xlsx', '.xls']:
                if file_extension == '.csv':
                    df = pd.read_csv(temp_path)
                else:
                    df = pd.read_excel(temp_path)
                st.dataframe(df, height=400)
            else:
                display_pdf(temp_path)
        
        with col2:
            st.subheader("Redacted Document")
            
            # Button to process the document
            if st.button("Start Redacting Sensitive Information"):
                job_id = str(uuid.uuid4())
                temp_dir = os.path.join("temp", job_id)
                os.makedirs(temp_dir, exist_ok=True)
                
                # Display progress bar
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                # Function to update progress
                def update_progress(progress):
                    progress_bar.progress(progress)
                    status_text.text(f"Processing: {int(progress * 100)}% complete")
                
                try:
                    # Process based on file type
                    start_time = time.time()
                    status_text.text("Processing document...")
                    
                    # Determine file type
                    file_extension = os.path.splitext(temp_path)[1].lower()
                    
                    # For Excel/CSV files, use Presidio only
                    if file_extension in ['.csv', '.xlsx', '.xls']:
                        # Load analyzer if not loaded already
                        _, _, analyzer = load_models()
                        
                        # Process the Excel/CSV file
                        progress_bar.progress(0.3)  # Show some progress
                        result = process_excel_csv(temp_path, analyzer)
                        process_time = time.time() - start_time
                        
                        # Display results
                        progress_bar.progress(1.0)
                        status_text.text(f"Processing completed in {process_time:.2f} seconds!")
                        
                        st.write(f"Found {result['sensitive_count']} sensitive items:")
                        if result['sensitive_count'] > 0:
                            items_to_show = min(10, len(result['sensitive_items']))
                            st.write(", ".join(result['sensitive_items'][:items_to_show]) + 
                                    ("..." if len(result['sensitive_items']) > items_to_show else ""))
                        
                        # Create download button for Excel
                        with open(result['excel_path'], "rb") as file:
                            st.download_button(
                                label="Download Processed Excel File",
                                data=file,
                                file_name=f"highlighted_{uploaded_file.name}",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                            
                        st.success(f"Sensitive information has been highlighted in the Excel file. You can download it using the button above.")
                    else:
                        # For PDF/images, use the existing process_document function
                        output_path = process_document(
                            temp_path, 
                            detection_method=selected_method,
                            llm_model=llm_model,
                            temp_dir=temp_dir,
                            progress_callback=update_progress
                        )
                        
                        # Display processed document
                        process_time = time.time() - start_time
                        status_text.text(f"Processing completed in {process_time:.2f} seconds!")
                        progress_bar.progress(1.0)
                        is_image_output = output_path.lower().endswith(('.jpg', '.jpeg', '.png', '.bmp', '.tiff'))
                        image_extension = os.path.splitext(output_path)[1].lower().replace('.', '')
                        mime_type = f"image/{image_extension}"
                        if is_image_output:
                            st.image(output_path, caption="Redacted Image")
                            
                            # Create download button for image
                            with open(output_path, "rb") as file:
                                st.download_button(
                                    label="Download Processed Image",
                                    data=file,
                                    file_name=f"redacted_{uploaded_file.name}",
                                    mime=mime_type
                                )
                        else:
                            display_pdf(output_path)
                        
                            # Create download button
                            with open(output_path, "rb") as file:
                                st.download_button(
                                    label="Download Processed PDF",
                                    data=file,
                                    file_name=f"redacted_{uploaded_file.name}",
                                    mime="application/pdf"
                                )
                    
                except Exception as e:
                    st.error(f"An error occurred during processing: {str(e)}")
                    st.exception(e)
    
    # Footer information
    st.markdown("---")
    st.markdown("""
    **Note:** This tool uses a combination of large language models (LLM) and syntactic analysis (Presidio) to detect sensitive information.
    However, no tool offers 100% accuracy. Always review the results to ensure all sensitive information has been redacted.
    """)

if __name__ == "__main__":
    main()