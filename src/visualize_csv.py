import pandas as pd
import re
import os
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

def highlight_sensitive_text(df, sensitive_text_list, sheet_name):
    """
    Creates an HTML file with highlighted sensitive text from the dataframe
    
    :param df: DataFrame containing the data
    :param sensitive_text_list: List of sensitive text to highlight
    :param sheet_name: Name of the sheet for the output filename
    :return: Path to the generated HTML file
    """
    # Create a styled DataFrame for HTML
    def highlight_entire_cell(s):
        is_sensitive = pd.Series(False, index=s.index)
        for col in s.index:
            if isinstance(s[col], str):
                for text in sensitive_text_list:
                    if text.lower() in str(s[col]).lower():
                        is_sensitive[col] = True
                        break
        
        return ['background-color: yellow; font-weight: bold' if is_sensitive[col] else '' for col in s.index]
    
    # Apply style to the DataFrame
    styled_df = df.style.apply(highlight_entire_cell, axis=1)
    
    # Create HTML with styling
    html_content = """
    <html>
    <head>
        <style>
            table {{
                border-collapse: collapse;
                width: 100%;
            }}
            th, td {{
                border: 1px solid #dddddd;
                text-align: left;
                padding: 8px;
            }}
            th {{
                background-color: #f2f2f2;
            }}
            tr:nth-child(even) td {{
                background-color: #f9f9f9;
            }}
            tr:nth-child(even) td[style*="background-color: yellow"] {{
                background-color: yellow !important;
            }}
        </style>
    </head>
    <body>
        <h1>Data from sheet "{sheet}" with Highlighted Sensitive Information</h1>
        {table}
    </body>
    </html>
    """.format(sheet=sheet_name, table=styled_df.to_html(escape=False, index=False))
    
    # Create output directory if it doesn't exist
    os.makedirs("../output", exist_ok=True)
    
    # Save the HTML file
    output_path = f"../output/highlighted_sensitive_data_{sheet_name}.html"
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html_content)
    
    return output_path
def save_to_excel_with_highlights(df, sensitive_text_list, output_path):
    """
    Saves dataframe to Excel with highlighted sensitive content
    
    :param df: DataFrame to save
    :param sensitive_text_list: List of sensitive text to highlight
    :param output_path: Path to save the Excel file
    :return: Path to the saved Excel file
    """
    # Save DataFrame to Excel
    df.to_excel(output_path, index=False)
    
    # Open the workbook and get active sheet
    workbook = openpyxl.load_workbook(output_path)
    worksheet = workbook.active
    
    # Define highlight fill
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Check each cell for sensitive text and apply highlighting
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        for col_idx, cell in enumerate(row, start=1):
            if cell.value and isinstance(cell.value, str):
                # Check if any sensitive text is in the cell
                for text in sensitive_text_list:
                    if text.lower() in cell.value.lower():
                        cell.fill = highlight_fill
                        break
    
    # Save the modified workbook
    workbook.save(output_path)
    return output_path
def save_to_excel_with_highlights(df, sensitive_text_list, sheet_name):
    """
    Saves dataframe to Excel with highlighted sensitive content
    
    :param df: DataFrame to save
    :param sensitive_text_list: List of sensitive text to highlight
    :param sheet_name: Name of the sheet for the output filename
    :return: Path to the saved Excel file
    """
    # Create output directory if it doesn't exist
    os.makedirs("../output", exist_ok=True)
    
    # Define output path
    output_path = f"../output/highlighted_sensitive_data_{sheet_name}.xlsx"
    
    # Save DataFrame to Excel
    df.to_excel(output_path, index=False)
    
    # Open the workbook and get active sheet
    workbook = openpyxl.load_workbook(output_path)
    worksheet = workbook.active
    
    # Define highlight fill
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Check each cell for sensitive text and apply highlighting
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        for col_idx, cell in enumerate(row, start=1):
            if cell.value and isinstance(cell.value, str):
                # Check if any sensitive text is in the cell
                for text in sensitive_text_list:
                    if text.lower() in cell.value.lower():
                        cell.fill = highlight_fill
                        break
    
    # Save the modified workbook
    workbook.save(output_path)
    return output_path

def visualize_sheet_data(df, sensitive_text_list, sheet_name):
    """
    Generate both HTML and Excel visualizations for the given dataframe
    
    :param df: DataFrame to visualize
    :param sensitive_text_list: List of sensitive text to highlight
    :param sheet_name: Name of the sheet being processed
    :return: Tuple of (html_path, excel_path)
    """
    # Sort sensitive text by length (descending) to ensure longer matches are processed first
    sorted_sensitive_text = sorted(sensitive_text_list, key=len, reverse=True)
    
    # Generate the highlighted HTML
    html_path = highlight_sensitive_text(df, sorted_sensitive_text, sheet_name)
    
    # Save to Excel with highlighting
    excel_path = save_to_excel_with_highlights(df, sorted_sensitive_text, sheet_name)
    
    return html_path, excel_path

# # Load the data
# df = pd.read_csv("Final Sec Attendee List.csv")

# # Load sensitive text list
# list_sensitive_text = open("pii_entities.txt").read().split("\n")
# list_sensitive_text = [text for text in list_sensitive_text if text != ""]

# # Sort sensitive text by length (descending) to ensure longer matches are processed first
# list_sensitive_text.sort(key=len, reverse=True)

# # Generate the highlighted HTML
# output_file = highlight_sensitive_text(df, list_sensitive_text)

# # Save to Excel with highlighting
# excel_output_path = "highlighted_sensitive_data.xlsx"
# excel_file = save_to_excel_with_highlights(df, list_sensitive_text, excel_output_path)

# print(f"Highlighted HTML saved to {os.path.abspath(output_file)}")
# print(f"Highlighted Excel data saved to {os.path.abspath(excel_file)}")
# print(f"Open these files to view the highlighted sensitive information.")